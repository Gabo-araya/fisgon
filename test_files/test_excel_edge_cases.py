#!/usr/bin/env python3
"""
Test script para casos edge de extracción de contenido Excel
"""

import sys
import os

# Agregar el directorio del proyecto al path
sys.path.append('/home/gabo/proy/django_proy/fisgon')

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
import django
django.setup()

from crawler.extractors import extract_full_content_from_file

def create_edge_case_xlsx():
    """Crear archivo XLSX con casos edge"""
    import openpyxl
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # Hoja 1: Hoja vacía
    ws1 = wb.active
    ws1.title = 'HojaVacia'
    # No agregar datos
    
    # Hoja 2: Solo una celda
    ws2 = wb.create_sheet('UnaCelda')
    ws2['A1'] = 'Solo una celda'
    
    # Hoja 3: Datos con fórmulas
    ws3 = wb.create_sheet('ConFormulas')
    ws3['A1'] = 'Número'
    ws3['B1'] = 'Fórmula'
    ws3['A2'] = 10
    ws3['B2'] = '=A2*2'
    
    # Hoja 4: Muchos datos (para probar límites)
    ws4 = wb.create_sheet('MuchosDatos')
    for i in range(200):  # Más de 100 filas (límite)
        for j in range(60):  # Más de 50 columnas (límite)
            ws4.cell(row=i+1, column=j+1, value=f'Celda_{i}_{j}')
    
    # Hoja 5: Caracteres especiales y Unicode
    ws5 = wb.create_sheet('Unicode')
    ws5['A1'] = 'Español: ñáéíóú'
    ws5['A2'] = 'Emoji: 🔍📊💾'
    ws5['A3'] = 'Símbolos: $#@%&*'
    
    file_path = 'test_files/test_excel_edge_cases.xlsx'
    wb.save(file_path)
    print(f'✅ Archivo edge cases creado: {file_path}')
    return file_path

def test_edge_cases():
    """Probar casos edge"""
    
    # Crear archivo de prueba
    test_file = create_edge_case_xlsx()
    
    print(f'\n=== PROBANDO CASOS EDGE: {test_file} ===')
    
    try:
        content = extract_full_content_from_file(test_file)
        
        if content:
            print(f'✅ Contenido extraído: {len(content)} caracteres')
            
            # Verificar que maneja hojas vacías
            if 'HojaVacia' in content:
                if '(Hoja vacía)' in content:
                    print('✅ Maneja hojas vacías correctamente')
                else:
                    print('⚠️ Hoja vacía procesada pero sin mensaje')
            
            # Verificar límites
            if '(limitado a' in content:
                print('✅ Aplica límites de celdas correctamente')
            else:
                print('ℹ️ No se alcanzó el límite de celdas')
            
            # Verificar Unicode
            if '🔍' in content or 'ñ' in content:
                print('✅ Maneja caracteres Unicode correctamente')
            else:
                print('⚠️ Posible problema con caracteres Unicode')
            
            # Mostrar estructura
            lines = content.split('\n')
            sheet_count = len([line for line in lines if line.startswith('=== HOJA:')])
            print(f'✅ Hojas procesadas: {sheet_count}')
            
            # Mostrar preview
            print('\n📄 Preview (primeras 20 líneas):')
            for i, line in enumerate(lines[:20]):
                print(f'   {i+1:2d}: {line}')
            
            if len(lines) > 20:
                print('   ...')
                
        else:
            print('❌ No se extrajo contenido')
            
    except Exception as e:
        print(f'❌ Error: {str(e)}')
        import traceback
        traceback.print_exc()

def test_performance():
    """Probar performance básica"""
    print('\n=== PRUEBA DE PERFORMANCE ===')
    
    import time
    files = [
        'test_files/test_excel.xlsx',
        'test_files/test_excel.xls',
        'test_files/test_excel_edge_cases.xlsx'
    ]
    
    for file_path in files:
        if os.path.exists(file_path):
            start_time = time.time()
            content = extract_full_content_from_file(file_path)
            end_time = time.time()
            
            duration = end_time - start_time
            chars_per_sec = len(content) / duration if duration > 0 else 0
            
            print(f'📊 {os.path.basename(file_path)}:')
            print(f'   Tiempo: {duration:.3f} segundos')
            print(f'   Contenido: {len(content)} caracteres')
            print(f'   Velocidad: {chars_per_sec:.0f} chars/seg')

if __name__ == "__main__":
    print("🔍 INICIANDO PRUEBAS DE CASOS EDGE")
    print("=" * 60)
    
    test_edge_cases()
    test_performance()
    
    print("\n" + "=" * 60)
    print("✅ PRUEBAS DE CASOS EDGE COMPLETADAS")